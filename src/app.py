import streamlit as st
import pandas as pd
import io
import sys
import tempfile
from load_data import load_and_process_data, create_nutrient_constraints, load_all_menus
from evaluation_function import calculate_harmony_matrix, get_top_n_harmony_pairs, validate_weekly_constraints, validate_weekly_constraints_detailed, calculate_actual_cost
from spea2_optimizer import SPEA2Optimizer
from Diet_class import NutrientConstraints, set_servings, get_servings
from diet_converter import convert_diet_format
from food_mapper import apply_food_mapping
import time
from datetime import datetime, timezone, timedelta
from utils import diet_to_dataframe, count_menu_changes
import random
import os
from github import Github, Auth
import base64
import uuid
from pathlib import Path

st.set_page_config(page_title="요양원 식단 최적화 프로그램", layout="wide")

KST = timezone(timedelta(hours=9))

# GitHub 토큰 자동 로드
def load_github_token():
    """Streamlit Secrets, 환경 변수 또는 .env 파일에서 GitHub 토큰 로드"""
    # 1. Streamlit Secrets에서 확인 (Streamlit Cloud용)
    try:
        if hasattr(st, 'secrets') and 'GITHUB_TOKEN' in st.secrets:
            return st.secrets['GITHUB_TOKEN']
    except Exception:
        pass
    
    # 2. 환경 변수에서 확인
    token = os.getenv('GITHUB_TOKEN')
    if token:
        return token
    
    # 3. .env 파일에서 확인
    env_path = Path(__file__).parent.parent / '.env'
    if env_path.exists():
        try:
            with open(env_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        if key.strip() == 'GITHUB_TOKEN':
                            return value.strip()
        except Exception:
            pass
    
    # 4. 기본 토큰 (평가용 - 실제 토큰은 .env 또는 Streamlit Secrets에 설정)
    return ''  # 빈 문자열 반환 시 사용자에게 입력 요청

# 기본 GitHub 토큰 설정
DEFAULT_GITHUB_TOKEN = load_github_token()

st.markdown("""
<style>
    .reportview-container {
        background: #f0f2f6
    }
    .main > div {
        padding-top: 2rem;
    }
    .stButton>button {
        background-color: #FAFAFA !important;
        color: black !important;
        font-size: 20px !important;
        font-weight: bold !important;
        padding: 15px 0 !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 10px 10px !important;
        border-radius: 8px !important;
        display: flex;
        justify-content: center;
        align-items: center;
        width: 100%;
    }
    .stButton:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    transition: all 0.3s ease;
    }
    .div[data-testid="element-container"]:has(button:contains("↻")) button {
        background-color: transparent !important;
        border: none !important;
        color: #2d2736 !important;
        padding: 0.rem !important;
        font-weight: bold;
    }
    .menu-item {
        background-color: #ffffff;
        border-radius: 10px;
        padding: 10px;
        margin-bottom: 10px;
    }
    [data-testid="stDataFrame"] tbody tr {
        height: 150px !important;
    }
    [data-testid="stDataFrame"] tbody td {
        vertical-align: top !important;
        padding: 15px !important;
        line-height: 1.8 !important;
        white-space: pre-wrap !important;
    }
</style>
""", unsafe_allow_html=True)

USERS = {
    "SR01": "test01",
    "SR02": "test02",
    "SR03": "test03",
    "SR04": "test04",
    "SR05": "test05",
    "SR06": "test06",
    "SR07": "test07",
    "SR08": "test08",
    "SR09": "test09",
    "SR010": "test10",
    "SR011": "test11",
    "SR012": "test12",
    "SR013": "test13",
}

if 'weekly_diet' not in st.session_state:
    st.session_state.weekly_diet = None
if 'initial_fitness' not in st.session_state:
    st.session_state.initial_fitness = None
if 'initial_cost' not in st.session_state:
    st.session_state.initial_cost = None
if 'nutrients_data' not in st.session_state:
    st.session_state.nutrients_data = None
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'username' not in st.session_state:
    st.session_state.username = ""
if 'file_uploaded' not in st.session_state:
    st.session_state.file_uploaded = False
if 'optimization_complete' not in st.session_state:
    st.session_state.optimization_complete = False
if 'optimization_results' not in st.session_state:
    st.session_state.optimization_results = {}
if 'optimization_start_time' not in st.session_state:
    st.session_state.optimization_start_time = None
if 'optimization_end_time' not in st.session_state:
    st.session_state.optimization_end_time = None
if 'optimization_duration' not in st.session_state:
    st.session_state.optimization_duration = None
if 'user_id' not in st.session_state:
    st.session_state.user_id = str(uuid.uuid4())
if 'user_servings' not in st.session_state:
    st.session_state.user_servings = 55
if 'github_token' not in st.session_state:
    st.session_state.github_token = DEFAULT_GITHUB_TOKEN

def get_user_servings():
    return st.session_state.user_servings

def set_user_servings(value):
    st.session_state.user_servings = value
    set_servings(value)

def login_page():
    st.markdown("""
    <style>
        .login-container {
            max-width: 400px;
            margin: 0 auto;
            padding: 2rem;
            border-radius: 10px;
            background-color: #f8f9fa;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }
        .login-title {
            text-align: center;
            color: #2d2736;
            margin-bottom: 2rem;
        }
    </style>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:        
        st.markdown('<h2 class="login-title">🔒 시스템 로그인</h2>', unsafe_allow_html=True)
        with st.form("login_form"):
            username = st.text_input("사용자명", placeholder="부여받은 아이디를 입력하세요")
            password = st.text_input("비밀번호", type="password", placeholder="부여받은 비밀번호를 입력하세요")
            submit_button = st.form_submit_button("로그인", use_container_width=True)
            if submit_button:
                if username in USERS and USERS[username] == password:
                    st.session_state.logged_in = True
                    st.session_state.username = username
                    st.experimental_rerun()
                else:
                    st.error("사용자명 또는 비밀번호가 올바르지 않습니다.")

def logout():
    st.session_state.logged_in = False
    st.session_state.username = ""
    st.experimental_rerun()

@st.cache_data
def load_data():
    name = 'jeongseong'
    diet_db_path = f'./data/sarang_DB/processed_DB/DIET_{name}.xlsx'
    menu_db_path = f'./data/sarang_DB/processed_DB/Menu_ingredient_nutrient_{name}.xlsx'
    ingre_db_path = f'./data/sarang_DB/processed_DB/Ingredient_Price_{name}.xlsx'

    diet_db = load_and_process_data(diet_db_path, menu_db_path, ingre_db_path)
    nutrient_constraints = create_nutrient_constraints()
    harmony_matrix, menus, menu_counts, _ = calculate_harmony_matrix(diet_db)
    all_menus = load_all_menus(menu_db_path, ingre_db_path)
    
    return diet_db, nutrient_constraints, harmony_matrix, menus, menu_counts, all_menus

def calculate_improvements(initial_fitness, optimized_fitness):
    improvements = []
    for init, opt in zip(initial_fitness, optimized_fitness):
        if init != 0:
            imp = (opt - init) / abs(init) * 100
        else:
            imp = float('inf') if opt > 0 else (0 if opt == 0 else float('-inf'))
        improvements.append(imp)
    return improvements

def handle_reupload():
    st.session_state.file_uploaded = False
    st.session_state.optimization_complete = False
    st.session_state.optimization_results = {}
    st.session_state.weekly_diet = None
    st.session_state.initial_fitness = None
    st.session_state.initial_cost = None
    st.session_state.nutrients_data = None
    st.session_state.optimization_start_time = None
    st.session_state.optimization_end_time = None
    st.session_state.optimization_duration = None
    if 'uploaded_file' in st.session_state:
        del st.session_state.uploaded_file
    if 'random_diet' in st.session_state:
        del st.session_state.random_diet

def generate_random_weekly_diet():
    diet_db_path = './data/sarang_DB/processed_DB/DIET_jeongseong.xlsx'
    df = pd.read_excel(diet_db_path)

    unique_days = df['Day'].unique()
    selected_days = random.sample(list(unique_days), 7)
    selected_days.sort()

    selected_meals = df[df['Day'].isin(selected_days)].copy()

    day_mapping = {day: i+1 for i, day in enumerate(selected_days)}
    selected_meals['Day'] = selected_meals['Day'].map(day_mapping)

    meal_order = {'Breakfast': 1, 'Lunch': 2, 'Dinner': 3}
    selected_meals['sort_order'] = selected_meals['MealType'].map(meal_order)
    selected_meals = selected_meals.sort_values(['Day', 'sort_order'])
    selected_meals = selected_meals.drop('sort_order', axis=1)

    return selected_meals

def process_mapped_diet_data(file_path):
    try:
        df = pd.read_excel(file_path)

        if 'Mapped_Menus' in df.columns:
            standard_df = df[['Day', 'MealType', 'Mapped_Menus']].copy()
            standard_df.rename(columns={'Mapped_Menus': 'Menus'}, inplace=True)

            user_id = st.session_state.user_id
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'_{user_id}_standard.xlsx') as tmp_file:
                standard_file_path = tmp_file.name

            standard_df.to_excel(standard_file_path, index=False)
            return standard_file_path
        else:
            return file_path

    except Exception as e:
        st.error(f"⏱ 매핑된 데이터 처리 중 오류 발생: {str(e)}")
        return file_path

def detect_and_convert_diet_format(uploaded_file):
    try:
        user_id = st.session_state.user_id
        with tempfile.NamedTemporaryFile(delete=False, suffix=f'_{user_id}.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getbuffer())
            temp_input_path = tmp_file.name

        df = pd.read_excel(temp_input_path)

        expected_columns = ['Day', 'MealType', 'Menus']
        if all(col in df.columns for col in expected_columns):
            mapping_file_path = './src/food_mapping.csv'

            with tempfile.NamedTemporaryFile(delete=False, suffix=f'_{user_id}_mapped.xlsx') as tmp_mapped:
                temp_mapped_path = tmp_mapped.name

            apply_food_mapping(temp_input_path, mapping_file_path, temp_mapped_path)
            os.unlink(temp_input_path)

            final_path = process_mapped_diet_data(temp_mapped_path)
            if final_path != temp_mapped_path:
                os.unlink(temp_mapped_path)

            return final_path

        if '주간 식단표' in df.columns or len(df.columns) >= 7:
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'_{user_id}_converted.xlsx') as tmp_converted:
                temp_output_path = tmp_converted.name

            convert_diet_format(temp_input_path, temp_output_path)
            os.unlink(temp_input_path)

            mapping_file_path = './src/food_mapping.csv'

            with tempfile.NamedTemporaryFile(delete=False, suffix=f'_{user_id}_mapped.xlsx') as tmp_mapped:
                temp_mapped_path = tmp_mapped.name

            apply_food_mapping(temp_output_path, mapping_file_path, temp_mapped_path)
            os.unlink(temp_output_path)
            final_path = process_mapped_diet_data(temp_mapped_path)
            if final_path != temp_mapped_path:
                os.unlink(temp_mapped_path)
            return final_path

        else:
            st.error("⏱ 지원되지 않는 파일 형태입니다.")
            os.unlink(temp_input_path)
            return None

    except Exception as e:
        st.error(f"⏱ 파일 처리 중 오류가 발생했습니다: {str(e)}")
        if 'temp_input_path' in locals():
            os.unlink(temp_input_path)
        return None

def create_weekly_diet_table(weekly_diet, title="주간 식단표", return_menu_counts=False):
    from datetime import datetime, timedelta

    days_data = {}
    meal_types = ['Breakfast', 'Lunch', 'Dinner']
    korean_meals = ['아침', '점심', '저녁']

    for day in range(1, 8):
        days_data[day] = {meal: [] for meal in meal_types}

    for i, meal in enumerate(weekly_diet.meals):
        day = (i // 3) + 1
        meal_type = meal_types[i % 3]

        if day <= 7:
            menu_names = [menu.name for menu in meal.menus]
            days_data[day][meal_type] = menu_names

    max_menus = {}
    for meal_type in meal_types:
        max_count = 0
        for day in range(1, 8):
            count = len(days_data[day][meal_type])
            if count > max_count:
                max_count = count
        max_menus[meal_type] = max_count

    today = datetime.now().date()
    weekdays = []
    for i in range(7):
        date = today + timedelta(days=i)
        weekdays.append(date)

    table_data = []

    header_row = ['구분'] + weekdays
    table_data.append(header_row)

    for meal_idx, (meal_type, korean_meal) in enumerate(zip(meal_types, korean_meals)):
        max_menu_count = max_menus[meal_type]

        for menu_idx in range(max_menu_count):
            if menu_idx == 0:
                row = [korean_meal]
            else:
                row = ['']

            for day in range(1, 8):
                menus = days_data[day][meal_type]
                if menu_idx < len(menus):
                    row.append(menus[menu_idx])
                else:
                    row.append('')

            table_data.append(row)

    if return_menu_counts:
        return pd.DataFrame(table_data), max_menus
    else:
        return pd.DataFrame(table_data)

def upload_to_github(file_buffer, filename, github_token=None, repo_name="diet-optimization-results"):
    if not github_token:
        github_token = st.session_state.get('github_token')

    if not github_token:
        return {
            'success': False,
            'error': 'GitHub 토큰이 설정되지 않았습니다.'
        }

    try:
        auth = Auth.Token(github_token)
        g = Github(auth=auth)
        user = g.get_user()
        repo = user.get_repo(repo_name)
        file_content = file_buffer.getvalue()

        today = datetime.now().strftime("%Y-%m-%d")
        file_path = f"results/{today}/{filename}"

        try:
            existing_file = repo.get_contents(file_path)
            repo.update_file(
                file_path,
                f"Update {filename}",
                file_content,
                existing_file.sha
            )
        except:
            repo.create_file(
                file_path,
                f"Add {filename}",
                file_content
            )

        return {
            'success': True,
            'repo_url': repo.html_url,
            'file_path': file_path
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'GitHub 업로드 실패: {str(e)}'
        }

def export_results_to_excel():
    if not st.session_state.optimization_complete or not st.session_state.optimization_results:
        return None

    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        improved_diets = st.session_state.optimization_results
        all_improvements = [improvements for _, _, improvements in improved_diets]
        avg_improvements = [sum(imp[i] for imp in all_improvements) / len(all_improvements) for i in range(4)]

        all_change_rates = []
        for diet, _, _ in improved_diets:
            menu_changes = count_menu_changes(st.session_state.weekly_diet, diet)
            total_menus = sum(counts['total'] for counts in menu_changes.values())
            total_changed = sum(counts['changed'] for counts in menu_changes.values())
            change_rate = (total_changed / total_menus * 100) if total_menus > 0 else 0
            all_change_rates.append(change_rate)
        overall_change_rate = sum(all_change_rates) / len(all_change_rates) if all_change_rates else 0

        summary_data = {
            "항목": ["사용자", "알고리즘", "세대수", "시작시간", "완료시간", "소요시간", "개선된 해 개수", "평균 영양 개선율(%)", "평균 비용 개선율(%)", "평균 조화 개선율(%)", "평균 다양성 개선율(%)", "평균 메뉴 변경율(%)"],
            "값": [
                st.session_state.username,
                "SPEA2",
                getattr(st.session_state, 'generations', 'N/A'),
                st.session_state.optimization_start_time.strftime("%Y-%m-%d %H:%M:%S") if st.session_state.optimization_start_time else 'N/A',
                st.session_state.optimization_end_time.strftime("%Y-%m-%d %H:%M:%S") if st.session_state.optimization_end_time else 'N/A',
                f"{st.session_state.optimization_duration:.1f}초" if st.session_state.optimization_duration else 'N/A',
                len(improved_diets),
                f"{avg_improvements[0]:.2f}" if len(avg_improvements) > 0 else 'N/A',
                f"{avg_improvements[1]:.2f}" if len(avg_improvements) > 1 else 'N/A',
                f"{avg_improvements[2]:.2f}" if len(avg_improvements) > 2 else 'N/A',
                f"{avg_improvements[3]:.2f}" if len(avg_improvements) > 3 else 'N/A',
                f"{overall_change_rate:.1f}"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='📊 최적화 요약', index=False)

        ws_summary = writer.sheets['📊 최적화 요약']
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 25

        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment

        for j, (optimized_diet, optimized_fitness, improvements) in enumerate(improved_diets):
            days = len(optimized_diet.meals) // 3
            current_servings = get_user_servings()
            optimized_cost = calculate_actual_cost(optimized_diet, current_servings)
            initial_cost = st.session_state.initial_cost
            cost_change = initial_cost - optimized_cost
            menu_changes = count_menu_changes(st.session_state.weekly_diet, optimized_diet)

            sheet_name = f'💡 제안식단 {j+1}'

            weekly_diet_table, max_menus_info = create_weekly_diet_table(optimized_diet, f"제안 식단 {j+1}", return_menu_counts=True)
            weekly_diet_table.to_excel(writer, sheet_name=sheet_name, startrow=1, index=False, header=False)

            ws = writer.sheets[sheet_name]
            ws['A1'] = f"🍽️ 최적화된 주간 식단표 - 제안 식단 {j+1}"
            ws['A1'].font = Font(bold=True, size=14, color="2F5597")

            nutrients_start = 22
            nutrients_data = []
            for nutrient in nutrient_constraints.min_values.keys():
                total = sum(sum(menu.get_adjusted_nutrients()[nutrient] for menu in meal.menus) for meal in optimized_diet.meals)
                daily_avg = total / days
                min_val = nutrient_constraints.min_values[nutrient]
                max_val = nutrient_constraints.max_values[nutrient]
                status = "✅" if min_val <= daily_avg <= max_val else "⚠️"

                nutrients_data.append({
                    "영양소": nutrient,
                    "일일평균": f"{daily_avg:.1f}",
                    "권장범위": f"{min_val} ~ {max_val}",
                    "상태": status
                })
            nutrients_df = pd.DataFrame(nutrients_data)
            nutrients_df.to_excel(writer, sheet_name=sheet_name, startrow=nutrients_start, index=False)

            ws[f'A{nutrients_start}'] = "🎯 영양성분 분석"
            ws[f'A{nutrients_start}'].font = Font(bold=True, size=14, color="2F5597")

            cost_start = 30
            cost_data = {
                "항목": ["총 식재료 비용", "1인당 비용", "1라당 비용"],
                "금액": [
                    f"{optimized_cost:,.0f}원",
                    f"{optimized_cost/current_servings:,.0f}원" if current_servings > 0 else "N/A",
                    f"{optimized_cost/(current_servings*21):,.0f}원" if current_servings > 0 else "N/A"
                ]
            }
            cost_df = pd.DataFrame(cost_data)
            cost_df.to_excel(writer, sheet_name=sheet_name, startrow=cost_start, index=False)
            
            ws[f'A{cost_start}'] = "💰 총 식재료 비용 정보"
            ws[f'A{cost_start}'].font = Font(bold=True, size=14, color="2F5597")

            perform_start = 36
            performance_data = {
                "지표": ["영양 점수", "비용 점수", "조화 점수", "다양성 점수", "총 식재료 비용(원)"],
                "초기값": [
                    f"{st.session_state.initial_fitness[0]:.2f}",
                    f"{st.session_state.initial_fitness[1]:.2f}",
                    f"{st.session_state.initial_fitness[2]:.2f}",
                    f"{st.session_state.initial_fitness[3]:.2f}",
                    f"{initial_cost:,.0f}",
                ],
                "최적화값": [
                    f"{optimized_fitness[0]:.2f}",
                    f"{optimized_fitness[1]:.2f}",
                    f"{optimized_fitness[2]:.2f}",
                    f"{optimized_fitness[3]:.2f}",
                    f"{optimized_cost:,.0f}",
                ],
                "개선율(%)": [
                    f"{improvements[0]:.2f}",
                    f"{improvements[1]:.2f}",
                    f"{improvements[2]:.2f}",
                    f"{improvements[3]:.2f}",
                    f"{cost_change:,.0f}원",
                ]
            }
            performance_df = pd.DataFrame(performance_data)
            performance_df.to_excel(writer, sheet_name=sheet_name, startrow=perform_start, index=False)

            ws[f'A{perform_start}'] = "📈 기존 식단과의 성능 비교"
            ws[f'A{perform_start}'].font = Font(bold=True, size=14, color="2F5597")

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width

    buffer.seek(0)
    return buffer

def parallel_optimize(optimizer, diet_db, weekly_diet, generations):
    """단순 최적화 실행"""
    return optimizer.optimize(diet_db, weekly_diet, generations)

if not st.session_state.logged_in:
    login_page()
    st.stop()

col1, col2, col3 = st.columns([3, 5, 1])
with col1:
    st.markdown(f'<p style="font-size: 20px; font-weight: normal; margin-top: 4px;">🙋‍♀️ {st.session_state.username}님 환영합니다!</p>', unsafe_allow_html=True)
with col3:
    if st.button("로그아웃", key="logout_btn"):
        logout()

diet_db, default_constraints, harmony_matrix, menus, menu_counts, all_menus = load_data()

with st.sidebar:
    col1, col2, col3 = st.columns([1, 5, 1])
    with col2:
        st.image("./assets/logo.png", width=180, use_column_width=True)

    st.markdown("---")
    st.subheader("📥 샘플 식단표 다운로드")
    
    # 식단표 A
    with open('./data/File_A.xlsx', 'rb') as f:
        st.download_button(
            label="식단표 A 다운로드",
            data=f,
            file_name="File_A.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    # 식단표 B
    with open('./data/File_B.xlsx', 'rb') as f:
        st.download_button(
            label="식단표 B 다운로드",
            data=f,
            file_name="File_B.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    st.markdown("---")
    st.subheader("🔑부여받은 키 설정")
    
    # 기본 토큰이 설정되어 있으면 표시
    if st.session_state.github_token and st.session_state.github_token != "":
        st.success("✅ 평가용 키가 자동으로 설정되었습니다.")
        st.caption("필요시 아래에서 다른 키로 변경할 수 있습니다.")
    else:
        st.warning("⚠️ GitHub 키가 설정되지 않았습니다. 아래에 입력해주세요.")
    
    github_token = st.text_input(
        "GitHub 토큰을 입력해주세요" if not st.session_state.github_token else "다른 키를 사용하려면 입력해주세요 (선택사항)",
        type="password",
        placeholder="ghp_xxxxxxxxxxxxxxxx",
        value="",
        help="평가용 키가 환경 변수에 설정되어 있으면 자동 적용됩니다. 없으면 여기에 입력하세요."
    )
    
    if github_token:
        st.session_state.github_token = github_token
        st.info("🔄 새로운 키가 적용되었습니다.")
    elif 'github_token' not in st.session_state or not st.session_state.github_token:
        st.session_state.github_token = DEFAULT_GITHUB_TOKEN
            
    st.markdown("---")
    st.subheader("🍽️ 조리 인분 설정")
    servings = st.number_input(
        "서빙 인원수",
        min_value=1,
        max_value=200,
        value=get_user_servings(),
        step=1,
        help="식단을 준비할 인원수를 설정합니다."
    )
    set_user_servings(servings)
    
    st.markdown("---")
    st.subheader("🔧 영양소 제한 설정")
    st.markdown('<p style="font-size: 0.9em; color: #666;">각 영양소의 최소값, 최대값, 가중치를 설정하세요.</p>', unsafe_allow_html=True)
    user_min_values = {}
    user_max_values = {}
    user_weights = {}
    tabs = st.tabs(list(default_constraints.min_values.keys()))
    for i, nutrient in enumerate(default_constraints.min_values.keys()):
        with tabs[i]:
            st.markdown(f"#### {nutrient} 설정")
            min_range = 0.0
            max_range = float(default_constraints.max_values[nutrient] * 2)
            default_min = float(default_constraints.min_values[nutrient])
            default_max = float(default_constraints.max_values[nutrient])
            default_weight = float(default_constraints.weights[nutrient])
            
            user_min_values[nutrient] = st.slider(
                "최소값",
                min_value=min_range,
                max_value=max_range,
                value=default_min,
                step=10.0,
                help=f"{nutrient}의 일일 최소 권장량을 설정합니다."
            )
            
            user_max_values[nutrient] = st.slider(
                "최대값",
                min_value=user_min_values[nutrient],
                max_value=max_range,
                value=default_max,
                step=10.0,
                help=f"{nutrient}의 일일 최대 권장량을 설정합니다."
            )
            
            user_weights[nutrient] = st.slider(
                "가중치",
                min_value=0.1,
                max_value=5.0,
                value=default_weight,
                step=0.1,
                help="해당 영양소의 중요도를 설정합니다."
            )
    nutrient_constraints = NutrientConstraints(
        min_values=user_min_values,
        max_values=user_max_values,
        weights=user_weights
    )

st.markdown("---")
st.title('식단 최적화 프로그램')
optimizer = SPEA2Optimizer(all_menus, nutrient_constraints, harmony_matrix)
st.markdown("---")

if not st.session_state.file_uploaded:
    st.subheader('📂 초기 식단 설정')
    
    col1, col2 = st.columns(2)
    with col1:
        uploaded_file = st.file_uploader("초기 식단 파일을 업로드하세요", type="xlsx")
    with col2:
        if st.button("🎲 랜덤 식단 생성", use_container_width=True):
            st.session_state.file_uploaded = True
            st.session_state.uploaded_file = None
            st.session_state.random_diet = True
            st.experimental_rerun()
    
    if uploaded_file is not None:
        st.session_state.file_uploaded = True
        st.session_state.uploaded_file = uploaded_file
        st.session_state.random_diet = False
        st.experimental_rerun()
else:
    col1, col2 = st.columns([15, 1])
    with col1:
        if hasattr(st.session_state, 'random_diet') and st.session_state.random_diet:
            st.subheader('🎲 랜덤 생성된 초기 식단')
        else:
            st.subheader('👀 업로드된 초기 식단')
    with col2:
        st.button("↻", key="reupload_button", on_click=handle_reupload, help="다른 식단을 설정합니다", type="secondary")

    if st.session_state.weekly_diet is None:
        name = 'jeongseong'
        menu_db_path = f'./data/sarang_DB/processed_DB/Menu_ingredient_nutrient_{name}.xlsx'
        ingre_db_path = f'./data/sarang_DB/processed_DB/Ingredient_Price_{name}.xlsx'
        
        if hasattr(st.session_state, 'random_diet') and st.session_state.random_diet:
            random_diet_df = generate_random_weekly_diet()
            user_id = st.session_state.user_id
            temp_file = f"temp_random_diet_{user_id}.xlsx"
            random_diet_df.to_excel(temp_file, index=False)
            st.session_state.weekly_diet = load_and_process_data(temp_file, menu_db_path, ingre_db_path)
            os.remove(temp_file)
        else:
            uploaded_file = st.session_state.uploaded_file
            converted_file_path = detect_and_convert_diet_format(uploaded_file)

            if converted_file_path:
                try:
                    debug_df = pd.read_excel(converted_file_path)
                    old_stdout = sys.stdout
                    sys.stdout = captured_output = io.StringIO()

                    try:
                        st.session_state.weekly_diet = load_and_process_data(converted_file_path, menu_db_path, ingre_db_path)
                    finally:
                        sys.stdout = old_stdout

                    captured_text = captured_output.getvalue()
                    if "Missing menus" in captured_text:
                        missing_lines = [line for line in captured_text.split('\n') if "Missing menus" in line]
                        st.warning(f"⚠️ 데이터베이스에서 찾을 수 없어 제외된 메뉴들:\n" + "\n".join(missing_lines[:10]))

                    os.unlink(converted_file_path)
                except Exception as e:
                    st.error(f"⏱ 식단 데이터 로드 중 오류가 발생했습니다: {str(e)}")
                    if os.path.exists(converted_file_path):
                        os.unlink(converted_file_path)
                    st.stop()
            else:
                st.error("⏱ 파일을 처리할 수 없습니다.")
                st.stop()
    
        st.session_state.initial_fitness = optimizer.fitness(diet_db, st.session_state.weekly_diet)
        
        current_servings = get_user_servings()
        st.session_state.initial_cost = calculate_actual_cost(st.session_state.weekly_diet, current_servings)
        days = len(st.session_state.weekly_diet.meals) // 3
    
        nutrients_data = []
        for nutrient in nutrient_constraints.min_values.keys():
            total = sum(sum(menu.get_adjusted_nutrients()[nutrient] for menu in meal.menus) for meal in st.session_state.weekly_diet.meals)
            daily_avg = total / days
            min_val = nutrient_constraints.min_values[nutrient]
            max_val = nutrient_constraints.max_values[nutrient]
            
            status = "✅" if min_val <= daily_avg <= max_val else "⚠️"
            
            nutrients_data.append({
                "영양소": nutrient,
                "일일평균": f"{daily_avg:.1f}",
                "권장범위": f"{min_val} ~ {max_val}",
                "상태": status
            })
        
        st.session_state.nutrients_data = nutrients_data

    weekly_diet = st.session_state.weekly_diet
    initial_fitness = st.session_state.initial_fitness
    initial_cost = st.session_state.initial_cost

    st.dataframe(diet_to_dataframe(weekly_diet, "Initial Diet"), use_container_width=True)
    st.subheader("📊 초기 식단 일일 평균 영양성분")
    st.table(pd.DataFrame(st.session_state.nutrients_data))
    
    st.info(f"현재 설정: **{get_user_servings()}인분**으로 계산됨 (사이드바에서 변경 가능)")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
            <h3 style="margin: 0; color: #404040; font-size: 18px;">🥗 영양 점수</h3>
            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{initial_fitness[0]:.2f}점</p>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
            <h3 style="margin: 0; color: #404040; font-size: 18px;">💰 비용 점수</h3>
            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{initial_fitness[1]:.2f}점</p>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
            <h3 style="margin: 0; color: #404040; font-size: 18px;">👍 조화 점수</h3>
            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{initial_fitness[2]:.2f}점</p>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
            <h3 style="margin: 0; color: #404040; font-size: 18px;">💡 다양성 점수</h3>
            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{initial_fitness[3]:.2f}점</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown(f"**총 식재료 비용**: {initial_cost:,.0f}원")

    st.markdown("---")
    
    generations = st.slider("최적화 세대수 설정", min_value=50, max_value=500, value=200, step=50, 
                           help="세대수가 높을수록 더 좋은 결과를 얻을 수 있지만 시간이 더 오래 걸립니다.")
    
    if st.button("🚀 식단 최적화 시작", key="optimize_button"):
        st.session_state.generations = generations
        st.session_state.optimization_start_time = datetime.now(KST)
        start_time_for_duration = time.time()

        with st.spinner(f'식단 최적화 진행 중... ({generations}세대)'):
            pareto_front = parallel_optimize(optimizer, diet_db, weekly_diet, generations)
            
            st.session_state.optimization_end_time = datetime.now(KST)
            optimization_duration = time.time() - start_time_for_duration
            st.session_state.optimization_duration = optimization_duration

            constraint_satisfied_diets = []
            constraint_violated_diets = []

            for optimized_diet in pareto_front:
                optimized_fitness = optimizer.fitness(diet_db, optimized_diet)
                improvements = calculate_improvements(initial_fitness, optimized_fitness)
                improved_count = sum(1 for imp in improvements if imp > 0)
                
                if improved_count >= 3:
                    is_valid = validate_weekly_constraints(optimized_diet, nutrient_constraints)
                    diet_info = (optimized_diet, optimized_fitness, improvements)
                    
                    if is_valid:
                        constraint_satisfied_diets.append(diet_info)
                    else:
                        constraint_violated_diets.append(diet_info)
            
            improved_diets = constraint_satisfied_diets[:5]
            if len(improved_diets) < 5:
                needed = 5 - len(improved_diets)
                improved_diets.extend(constraint_violated_diets[:needed])
            
            st.session_state.optimization_results = improved_diets
            st.session_state.optimization_complete = True

    if st.session_state.optimization_complete and st.session_state.optimization_results:
        st.subheader('🏆 최적화 된 식단 🏆')
        st.caption('⭐️: 가장 많이 개선된 식단이에요!')
        improved_diets = st.session_state.optimization_results
        if improved_diets:
            total_improvements = []
            for _, _, improvements in improved_diets:
                total_improvement = sum(improvements) / len(improvements)
                total_improvements.append(total_improvement)

            best_diet_index = total_improvements.index(max(total_improvements))

            tab_names = []
            for i in range(len(improved_diets)):
                if i == best_diet_index:
                    tab_names.append(f"⭐ 제안 식단 {i+1}")
                else:
                    tab_names.append(f"제안 식단 {i+1}")

            diet_tabs = st.tabs(tab_names)
            
            for j, (diet_tab, (optimized_diet, optimized_fitness, improvements)) in enumerate(zip(diet_tabs, improved_diets)):
                with diet_tab:
                    st.dataframe(diet_to_dataframe(optimized_diet, f"SPEA2 - 제안 식단 {j+1}"), use_container_width=True)
                    
                    days = len(optimized_diet.meals) // 3
                    st.subheader("📊 일일 평균 영양성분")
                    nutrients_data = []
                    for nutrient in nutrient_constraints.min_values.keys():
                        total = sum(sum(menu.get_adjusted_nutrients()[nutrient] for menu in meal.menus) for meal in optimized_diet.meals)
                        daily_avg = total / days
                        min_val = nutrient_constraints.min_values[nutrient]
                        max_val = nutrient_constraints.max_values[nutrient]
                        
                        status = "✅" if min_val <= daily_avg <= max_val else "⚠️"
                        
                        nutrients_data.append({
                            "영양소": nutrient,
                            "일일평균": f"{daily_avg:.1f}",
                            "권장범위": f"{min_val} ~ {max_val}",
                            "상태": status
                        })
                    st.table(pd.DataFrame(nutrients_data))
                    
                    current_servings = get_user_servings()
                    optimized_cost = calculate_actual_cost(optimized_diet, current_servings)
                    cost_change = initial_cost - optimized_cost
                    
                    col1, col2, col3, col4 = st.columns(4)                    
                    improvement_colors = ['green' if imp > 0 else 'red' for imp in improvements]
                    with col1:
                        st.markdown(f"""
                        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
                            <h3 style="margin: 0; color: #404040; font-size: 18px;">🥗 영양 점수</h3>
                            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{optimized_fitness[0]:.2f} <span style="color: {improvement_colors[0]};">({improvements[0]:+.2f}%)</span></p>
                        </div>
                        """, unsafe_allow_html=True)
                    with col2:
                        st.markdown(f"""
                        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
                            <h3 style="margin: 0; color: #404040; font-size: 18px;">💰 비용 점수</h3>
                            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{optimized_fitness[1]:.2f} <span style="color: {improvement_colors[1]};">({improvements[1]:+.2f}%)</span></p>
                        </div>
                        """, unsafe_allow_html=True)
                    with col3:
                        st.markdown(f"""
                        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
                            <h3 style="margin: 0; color: #404040; font-size: 18px;">👍 조화 점수</h3>
                            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{optimized_fitness[2]:.2f} <span style="color: {improvement_colors[2]};">({improvements[2]:+.2f}%)</span></p>
                        </div>
                        """, unsafe_allow_html=True)
                    with col4:
                        st.markdown(f"""
                        <div style="padding: 10px; border-radius: 10px; background-color: #FAFAFA;">
                            <h3 style="margin: 0; color: #404040; font-size: 18px;">💡 다양성 점수</h3>
                            <p style="font-size: 20px; font-weight: bold; margin: 0px 0 0 0;">{optimized_fitness[3]:.2f} <span style="color: {improvement_colors[3]};">({improvements[3]:+.2f}%)</span></p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    st.markdown(f"**총 식재료 비용**: {optimized_cost:,.0f}원 ({cost_change:,.0f}원)")
                    
                    menu_changes = count_menu_changes(weekly_diet, optimized_diet)
                    st.markdown('---')
                    st.markdown('#### 📈 카테고리별 메뉴 변경 비율')
                    cols = st.columns(5)
                    for i, (category, counts) in enumerate(menu_changes.items()):
                        percentage = (counts['changed'] / counts['total']) * 100 if counts['total'] > 0 else 0
                        if percentage > 50:
                            color = "#FF7043"
                        elif percentage > 25:
                            color = "#FFB74D"
                        else:
                            color = "#81C784"
                        bar_width = percentage
                        cols[i % 5].markdown(f"""
                        <div style="margin-bottom: 15px; padding: 12px; border-radius: 8px; background-color: #F5F5F5;">
                            <div style="font-weight: bold; margin-bottom: 5px; font-size: 15px;">{category}</div>
                            <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                                <span>{counts['changed']}/{counts['total']}</span>
                                <span style="font-weight: bold;">{percentage:.1f}%</span>
                            </div>
                            <div style="background-color: #E0E0E0; height: 8px; border-radius: 4px; overflow: hidden;">
                                <div style="background-color: {color}; width: {bar_width}%; height: 100%;"></div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
            
            if (st.session_state.optimization_start_time and 
                st.session_state.optimization_end_time and 
                st.session_state.optimization_duration):
                
                st.markdown("---")
                st.subheader("📋 최적화 요약 정보")
                
                all_improvements = [improvements for _, _, improvements in improved_diets]
                avg_improvements = [sum(imp[i] for imp in all_improvements) / len(all_improvements) for i in range(4)]
                
                all_change_rates = []
                for diet, _, _ in improved_diets:
                    menu_changes = count_menu_changes(weekly_diet, diet)
                    total_menus = sum(counts['total'] for counts in menu_changes.values())
                    total_changed = sum(counts['changed'] for counts in menu_changes.values())
                    change_rate = (total_changed / total_menus * 100) if total_menus > 0 else 0
                    all_change_rates.append(change_rate)
                overall_change_rate = sum(all_change_rates) / len(all_change_rates) if all_change_rates else 0
                
                summary_data = {
                    "사용자": [st.session_state.username],
                    "알고리즘": ["SPEA2"],
                    "세대수": [generations],
                    "소요시간": [f"{st.session_state.optimization_duration:.1f}초"],
                    "개선된 해 개수": [len(improved_diets)],
                    "평균 개선율": [f"영양: {avg_improvements[0]:+.1f}% | 비용: {avg_improvements[1]:+.1f}% | 조화: {avg_improvements[2]:+.1f}% | 다양성: {avg_improvements[3]:+.1f}%"],
                    "평균 메뉴 변경율": [f"{overall_change_rate:.1f}%"]
                }
                
                summary_df = pd.DataFrame(summary_data).set_index("사용자")
                st.dataframe(summary_df, use_container_width=True)
                
                st.markdown("---")
                excel_buffer = export_results_to_excel()
                if excel_buffer:
                    unique_id = str(uuid.uuid4())[:8]
                    filename = f"식단최적화결과_{st.session_state.username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{unique_id}.xlsx"

                    col1, col2, col3 = st.columns([1, 1, 1])

                    with col1:
                        st.download_button(
                            label="📥 엑셀 다운로드",
                            data=excel_buffer,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                    with col2:
                        if st.button("파일 업로드", use_container_width=True):
                            with st.spinner('파일 업로드 중...'):
                                excel_buffer.seek(0)
                                result = upload_to_github(excel_buffer, filename)

                                if result and result.get('success'):
                                    st.success("✅ 파일 업로드 완료!")
                                else:
                                    error_msg = result.get('error', '알 수 없는 오류가 발생했습니다.')
                                    st.error(f"⏱ 파일 업로드 실패")
                                    if 'GitHub 토큰' in error_msg:
                                        st.info("💡 사이드바 부여받은 키 설정에 키를 입력해주세요.")

                    with col3:
                        st.markdown('''
                            <a href="https://docs.google.com/forms/d/e/1FAIpQLScM4wJH1PmNhNAgaOvtawqvQ3KgOQMXEjErl7KzodSOJGfU4w/viewform" target="_blank">
                                <button style="width:100%; padding:0.5rem; border:none;">사용성 평가</button>
                            </a>
                            ''', unsafe_allow_html=True)
                        st.caption('5주차에만 진행해주세요!')

        else:
            st.warning("3가지 이상 개선된 식단을 찾지 못했습니다. 다른 초기 식단으로 시도해보세요.")

        if st.button('🔄 새로운 최적화 실행'):
            st.session_state.optimization_complete = False
            st.session_state.optimization_results = {}
            st.experimental_rerun()

st.markdown("---")
st.caption("© 2025 요양원 식단 최적화 프로그램. All rights reserved.")
