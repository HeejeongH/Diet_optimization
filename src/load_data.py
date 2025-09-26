import pandas as pd
from typing import List, Dict, Tuple
from Diet_class import Ingredient, Menu, Meal, Diet, NutrientConstraints
from openpyxl import load_workbook

def _load_excel_files(menu_db_path: str, ingre_db_path: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    menu_ingre_df = pd.read_excel(menu_db_path, sheet_name='ingredient')
    menu_nutri_df = pd.read_excel(menu_db_path, sheet_name='nutrient')
    menu_cat_df = pd.read_excel(menu_db_path, sheet_name='category')
    
    ingre_price_df = pd.read_excel(ingre_db_path)
    ingre_price_df['단가(원/g)'] = pd.to_numeric(ingre_price_df['단가(원/g)'], errors='coerce')
    ingre_price_df['용량(g)'] = pd.to_numeric(ingre_price_df['용량(g)'], errors='coerce')
    
    return menu_ingre_df, menu_nutri_df, menu_cat_df, ingre_price_df

def _create_ingredient_dict(menu_ingre_df: pd.DataFrame, ingre_price_df: pd.DataFrame) -> Dict[str, List[Ingredient]]:
    ingredient_dict = {}
    
    for _, row in menu_ingre_df.iterrows():
        menu_name = row['Menu']
        ingredient_name = row['Ingredient']
        amount = pd.to_numeric(row['Amount_g'], errors='coerce')
        
        if pd.isna(amount):
            amount = 0.0
            print(f"Warning: Missing amount data for {ingredient_name}")
        
        price_info = ingre_price_df[ingre_price_df['Ingredient'] == ingredient_name]
        if not price_info.empty:
            price_per_g = price_info['단가(원/g)'].values[0]
            package_size = price_info['용량(g)'].values[0]
            
            if pd.isna(price_per_g) or pd.isna(package_size):
                price_per_g = 0.0
                package_size = 1.0
                print(f"Warning: Missing price/package data for {ingredient_name}")
        else:
            price_per_g = 0.0
            package_size = 1.0
            print(f"Warning: Missing price data for {ingredient_name}")
        
        if menu_name not in ingredient_dict:
            ingredient_dict[menu_name] = []
        
        ingredient_dict[menu_name].append(
            Ingredient(
                name=ingredient_name, 
                price_per_g=price_per_g, 
                amount_g=amount,
                package_size=package_size
            )
        )
    
    return ingredient_dict

def _create_menu_objects(menu_nutri_df: pd.DataFrame, ingredient_dict: Dict[str, List[Ingredient]], menu_categories: Dict[str, str]) -> Dict[str, Menu]:
    menu_objects = {}
    
    for _, row in menu_nutri_df.iterrows():
        menu_name = row['Menu']
        nutrients = {
            '에너지(kcal)': float(row['에너지(kcal)']),
            '탄수화물(g)': float(row['탄수화물(g)']),
            '단백질(g)': float(row['단백질(g)']),
            '지방(g)': float(row['지방(g)']),
            '식이섬유(g)': float(row['식이섬유(g)']),
        }
        ingredients = ingredient_dict.get(menu_name, [])
        category = menu_categories.get(menu_name, "Unknown")
        
        menu_objects[menu_name] = Menu(
            name=menu_name, 
            nutrients=nutrients, 
            ingredients=ingredients, 
            category=category
        )
    
    return menu_objects

def _normalize_menu_name(menu_name: str, available_menus: list) -> str:
    """메뉴명 정규화 및 매칭"""
    menu_name = menu_name.strip()

    # 직접 매칭
    if menu_name in available_menus:
        return menu_name

    # 일반적인 매핑 규칙
    menu_mappings = {
        '흰밥': '쌀밥',
        '유부미소국': '배추유부미소국',
        '생선까스&타르소스': '생선까스',
        '우유계란찜': '계란찜',
        '돈육고추장볶음': '돼지고기고추장볶음',
        '올방개묵&양념장': '올방개묵',
        '두부구이*양념장': '두부구이',
        '납작만두&양념장': '납작만두',
        '브로콜리&초장': '브로콜리무침',
        # 추가 매핑
        '류산슬덮밥소스': '류산슬',
        '칼집후랑크야채볶음': '후랑크소세지볶음',
        '깻잎닭갈비': '닭갈비',
        '온도토리묵국': '도토리묵냉국',
        '호박새우젓국': '애호박새우젓국',
        '임연수소금구이': '임연수구이',
        '우민찌두부조림': '두부조림',
        '건파래볶음': '마른파래볶음',
        '돈삼겹수육': '수육',
        '비빔메밀국수': '간장비빔국수',
        '맑은김칫국': '김칫국',
        '숯불함박조림': '함박스테이크',
        '단호박팥찜': '단호박두부찜',
        '시금치국': '시금치된장국',
        '계란파국': '계란국',
        '맑은미역국': '미역국',
        '돈채고추잡채': '돼지고기채소볶음',
        '토마토스크램블': '스크램블에그',
        '안동찜닭': '찜닭',
        '맛살양배추볶음': '맛살볶음',
        '훈제오리야채볶음': '오리야채볶음',
        '둥근오이무침': '오이무침',
        '도시락김': '김자반',
        '너비아니야채볶음': '너비아니볶음',
        '감자채볶음': '감자볶음',
        '소고기탕국': '소고기무국',
        '콜라비무침': '콜라비나물',
        '미트볼데리야끼조림': '미트볼조림',
        '아주까리나물볶음': '아주까리나물',
        '뿌리채소영양밥': '영양밥',
        '쑥갓무생채': '쑥갓나물',
        '맑은배추국': '배추국',
        '방어조림': '방어구이',
        '쥬키니새우젓볶음': '주키니볶음'
    }

    if menu_name in menu_mappings:
        mapped_name = menu_mappings[menu_name]
        if mapped_name in available_menus:
            return mapped_name

    # 부분 매칭 시도 (앞부분 우선)
    for available_menu in available_menus:
        if menu_name.startswith(available_menu) or available_menu.startswith(menu_name):
            return available_menu

    # 키워드 매칭
    for available_menu in available_menus:
        if any(word in available_menu for word in menu_name.split('&') if len(word) > 1):
            return available_menu

    return None

def load_and_process_data(diet_db_path: str, menu_db_path: str, ingre_db_path: str) -> Diet:
    diet_df = pd.read_excel(diet_db_path)
    menu_ingre_df, menu_nutri_df, menu_cat_df, ingre_price_df = _load_excel_files(menu_db_path, ingre_db_path)
    
    menu_categories = dict(zip(menu_cat_df['Menu'], menu_cat_df['Category']))
    ingredient_dict = _create_ingredient_dict(menu_ingre_df, ingre_price_df)
    menu_objects = _create_menu_objects(menu_nutri_df, ingredient_dict, menu_categories)

    available_menu_names = list(menu_objects.keys())
    meals = []
    for _, row in diet_df.iterrows():
        meal_menus = row['Menus'].split(',')
        meal_menu_objects = []
        missing_menus = []
        mapped_menus = []

        for menu_name in meal_menus:
            original_name = menu_name.strip()
            normalized_name = _normalize_menu_name(original_name, available_menu_names)

            if normalized_name and normalized_name in menu_objects:
                meal_menu_objects.append(menu_objects[normalized_name])
                if normalized_name != original_name:
                    mapped_menus.append(f"{original_name} → {normalized_name}")
            else:
                missing_menus.append(original_name)

        if mapped_menus:
            print(f"Info: Menu mappings for Day {row['Day']} {row['MealType']}: {mapped_menus}")

        if missing_menus:
            print(f"Warning: Missing menus for Day {row['Day']} {row['MealType']}: {missing_menus}")

        meals.append(Meal(meal_menu_objects, str(row['Day']), row['MealType']))

    return Diet(meals)

def load_all_menus(menu_db_path: str, ingre_db_path: str) -> List[Menu]:
    menu_ingre_df, menu_nutri_df, menu_cat_df, ingre_price_df = _load_excel_files(menu_db_path, ingre_db_path)
    
    menu_categories = dict(zip(menu_cat_df['Menu'], menu_cat_df['Category']))
    ingredient_dict = _create_ingredient_dict(menu_ingre_df, ingre_price_df)
    menu_objects = _create_menu_objects(menu_nutri_df, ingredient_dict, menu_categories)
    
    return list(menu_objects.values())

def create_nutrient_constraints() -> NutrientConstraints:
    min_values = {
        '에너지(kcal)': 1440,
        '탄수화물(g)': 234,
        '단백질(g)': 54,
        '지방(g)': 32,
        '식이섬유(g)': 20,
    }
    max_values = {
        '에너지(kcal)': 2200,
        '탄수화물(g)': 357.5,
        '단백질(g)': 82.5,
        '지방(g)': 60,
        '식이섬유(g)': 50,
    }
    weights = {
        '에너지(kcal)': 1.0,
        '탄수화물(g)': 0.8,
        '단백질(g)': 1.4,
        '지방(g)': 1.2,
        '식이섬유(g)': 1.1,
    }
    return NutrientConstraints(min_values=min_values, max_values=max_values, weights=weights)

def load_sample_file(sample_path: str) -> None:
    sample = pd.read_excel(sample_path).iloc[4:22, 2:8]

    def extract_main_dish(dish: str) -> str:
        return dish.split('/')[0] if '/' in dish else dish

    result = []
    meal_types = ['Breakfast', 'Lunch', 'Dinner']
    
    for day in range(6):
        for idx, meal_type in enumerate(meal_types):
            start_idx = idx * 6
            menu_items = sample.iloc[start_idx:start_idx + 6, day].tolist()
            menu_items[0] = extract_main_dish(menu_items[0])
            menu_str = ', '.join(menu_items)
            result.append({
                'Day': day + 1,
                'MealType': meal_type,
                'Menus': menu_str
            })

    df_result = pd.DataFrame(result)

    book = load_workbook(sample_path)
    new_sheet_name = 'sample'
    sheet_number = 1
    while new_sheet_name in book.sheetnames:
        sheet_number += 1
        new_sheet_name = f'sample_{sheet_number}'

    with pd.ExcelWriter(sample_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        df_result.to_excel(writer, sheet_name=new_sheet_name, index=False)