import numpy as np
import pandas as pd
from typing import List, Dict, Any
from Diet_class import Diet
from optimizer_base import DietOptimizer
from scipy import stats
import pickle
import os
from openpyxl import Workbook

class PerformanceEvaluator:
    def __init__(self, diet_db: Diet, initial_diet: Diet, optimizers: Dict[str, DietOptimizer]):
        self.diet_db = diet_db
        self.initial_diet = initial_diet
        self.optimizers = optimizers
        self.metrics = ['hypervolume', 'spacing', 'diversity', 'convergence', 'execution_time']

    def calculate_hypervolume(self, solutions: List[Diet], optimizer: DietOptimizer) -> float:
        if not solutions:
            return 0.0
        reference_point = np.array([-100, -100, 0, 0])
        fitnesses = np.array([optimizer.fitness(self.diet_db, solution) for solution in solutions])
        if len(fitnesses) == 0:
            return 0.0
        normalized_fitnesses = (fitnesses - reference_point) / (np.array([0, 0, 100, 100]) - reference_point)
        sorted_points = normalized_fitnesses[np.lexsort(normalized_fitnesses.T)]
        volume = 0
        for i in range(len(sorted_points)):
            if i == 0:
                volume += np.prod(sorted_points[i])
            else:
                volume += np.prod(np.maximum(0, sorted_points[i] - sorted_points[i-1]))
        return volume

    def calculate_spacing(self, solutions: List[Diet], optimizer: DietOptimizer) -> float:
        if len(solutions) < 2:
            return 0.0
        fitnesses = np.array([optimizer.fitness(self.diet_db, solution) for solution in solutions])
        distances = []
        for i in range(len(fitnesses)):
            min_dist = float('inf')
            for j in range(len(fitnesses)):
                if i != j:
                    dist = np.linalg.norm(fitnesses[i] - fitnesses[j])
                    min_dist = min(min_dist, dist)
            distances.append(min_dist)
        mean_dist = np.mean(distances)
        return np.sqrt(np.sum((distances - mean_dist)**2) / (len(distances) - 1))

    def calculate_diversity(self, solutions: List[Diet], optimizer: DietOptimizer) -> float:
        if len(solutions) < 2:
            return 0.0
        fitnesses = np.array([optimizer.fitness(self.diet_db, solution) for solution in solutions])
        max_spread = np.max(fitnesses, axis=0) - np.min(fitnesses, axis=0)
        return np.mean(max_spread)

    def calculate_convergence(self, solutions: List[Diet], optimizer: DietOptimizer) -> float:
        if not solutions:
            return 0.0
        initial_fitness = np.array(optimizer.fitness(self.diet_db, self.initial_diet))
        final_fitnesses = np.array([optimizer.fitness(self.diet_db, solution) for solution in solutions])
        if len(final_fitnesses) == 0:
            return 0.0
        improvements = np.maximum(0, final_fitnesses - initial_fitness)
        return np.mean(improvements)

    def run_single_optimizer(self, optimizer_name: str, generations: int = 100, num_runs: int = 10, save_path: str = None):
        if optimizer_name not in self.optimizers:
            raise ValueError(f"Optimizer {optimizer_name} not found")
        
        optimizer = self.optimizers[optimizer_name]
        results = {metric: [] for metric in self.metrics}
        
        print(f"\nEvaluating {optimizer_name}...")
        
        for run in range(num_runs):
            import time
            start_time = time.time()
            
            try:
                solutions = optimizer.optimize(self.diet_db, self.initial_diet, generations)
                execution_time = time.time() - start_time
                
                if not solutions:
                    print(f"Warning: {optimizer_name} Run {run + 1} returned no solutions")
                    results['hypervolume'].append(0.0)
                    results['spacing'].append(0.0)
                    results['diversity'].append(0.0)
                    results['convergence'].append(0.0)
                else:
                    results['hypervolume'].append(self.calculate_hypervolume(solutions, optimizer))
                    results['spacing'].append(self.calculate_spacing(solutions, optimizer))
                    results['diversity'].append(self.calculate_diversity(solutions, optimizer))
                    results['convergence'].append(self.calculate_convergence(solutions, optimizer))
                
                results['execution_time'].append(execution_time)
                
            except Exception as e:
                print(f"Error in {optimizer_name} Run {run + 1}: {e}")
                for metric in self.metrics:
                    results[metric].append(0.0)
            
            print(f"Run {run + 1}/{num_runs} completed")
        
        if save_path:
            self.save_single_result(optimizer_name, results, save_path)
        
        return results

    def save_single_result(self, optimizer_name: str, results: Dict, save_path: str):
        os.makedirs(save_path, exist_ok=True)
        filename = os.path.join(save_path, f"{optimizer_name}_results.pkl")
        with open(filename, 'wb') as f:
            pickle.dump({optimizer_name: results}, f)
        print(f"Results saved to {filename}")

    def load_single_result(self, optimizer_name: str, save_path: str) -> Dict:
        filename = os.path.join(save_path, f"{optimizer_name}_results.pkl")
        with open(filename, 'rb') as f:
            return pickle.load(f)

    def combine_results(self, save_path: str, optimizer_names: List[str] = None) -> Dict:
        if optimizer_names is None:
            optimizer_names = list(self.optimizers.keys())
        
        combined_results = {}
        for name in optimizer_names:
            try:
                single_result = self.load_single_result(name, save_path)
                combined_results.update(single_result)
                print(f"Loaded results for {name}")
            except FileNotFoundError:
                print(f"Results file for {name} not found")
        
        return combined_results

    def perform_statistical_analysis(self, results: Dict[str, Dict[str, List[float]]]) -> Dict:
        statistical_results = {}
        
        for metric in self.metrics:
            statistical_results[metric] = {
                'normality': {},
                'overall_test': {},
                'pairwise_tests': {}
            }
            
            algorithm_data = {name: results[name][metric] for name in results.keys()}
            
            # Normality test
            for alg_name, data in algorithm_data.items():
                if len(data) > 2:
                    statistic, p_value = stats.shapiro(data)
                    statistical_results[metric]['normality'][alg_name] = {
                        'statistic': statistic,
                        'p_value': p_value
                    }
                else:
                    statistical_results[metric]['normality'][alg_name] = {
                        'statistic': 0.0,
                        'p_value': 1.0
                    }
            
            # Overall test
            all_data = [data for data in algorithm_data.values() if len(data) > 0]
            if len(all_data) > 1:
                try:
                    h_statistic, p_value = stats.kruskal(*all_data)
                    statistical_results[metric]['overall_test'] = {
                        'test': 'Kruskal-Wallis',
                        'statistic': h_statistic,
                        'p_value': p_value
                    }
                except ValueError:
                    statistical_results[metric]['overall_test'] = {
                        'test': 'Kruskal-Wallis',
                        'statistic': 0.0,
                        'p_value': 1.0
                    }
            
            # Pairwise comparisons
            alg_names = list(results.keys())
            for i in range(len(alg_names)):
                for j in range(i+1, len(alg_names)):
                    alg1, alg2 = alg_names[i], alg_names[j]
                    if len(algorithm_data[alg1]) > 0 and len(algorithm_data[alg2]) > 0:
                        try:
                            statistic, p_value = stats.mannwhitneyu(
                                algorithm_data[alg1],
                                algorithm_data[alg2],
                                alternative='two-sided'
                            )
                            statistical_results[metric]['pairwise_tests'][f'{alg1} vs {alg2}'] = {
                                'statistic': statistic,
                                'p_value': p_value
                            }
                        except ValueError:
                            statistical_results[metric]['pairwise_tests'][f'{alg1} vs {alg2}'] = {
                                'statistic': 0.0,
                                'p_value': 1.0
                            }
        
        return statistical_results

    def save_combined_results_to_excel(self, results: Dict, filename: str = 'combined_optimization_results.xlsx'):
        statistics = {}
        for name in results.keys():
            statistics[name] = {
                metric: {
                    'mean': np.mean(results[name][metric]),
                    'std': np.std(results[name][metric]),
                    'min': np.min(results[name][metric]),
                    'max': np.max(results[name][metric])
                }
                for metric in self.metrics
            }

        # Perform statistical analysis
        statistical_analysis = self.perform_statistical_analysis(results)

        # Save to Excel
        wb = Workbook()
        
        # Raw Results Sheet
        ws_raw = wb.active
        ws_raw.title = 'Raw Results'
        
        row = 1
        for metric in self.metrics:
            ws_raw.cell(row=row, column=1, value=metric.upper())
            row += 1
            
            headers = ['Algorithm'] + [f'Run {i+1}' for i in range(len(next(iter(results.values()))[metric]))]
            for col, header in enumerate(headers, 1):
                ws_raw.cell(row=row, column=col, value=header)
            
            for alg_name in results.keys():
                row += 1
                ws_raw.cell(row=row, column=1, value=alg_name)
                for col, value in enumerate(results[alg_name][metric], 2):
                    ws_raw.cell(row=row, column=col, value=value)
            
            row += 2

        # Summary Statistics Sheet
        ws_summary = wb.create_sheet('Summary Statistics')
        
        row = 1
        for metric in self.metrics:
            ws_summary.cell(row=row, column=1, value=metric.upper())
            row += 1
            
            headers = ['Algorithm', 'Mean', 'Std', 'Min', 'Max']
            for col, header in enumerate(headers, 1):
                ws_summary.cell(row=row, column=col, value=header)
            
            for alg_name in results.keys():
                row += 1
                stats = statistics[alg_name][metric]
                ws_summary.cell(row=row, column=1, value=alg_name)
                ws_summary.cell(row=row, column=2, value=stats['mean'])
                ws_summary.cell(row=row, column=3, value=stats['std'])
                ws_summary.cell(row=row, column=4, value=stats['min'])
                ws_summary.cell(row=row, column=5, value=stats['max'])
            
            row += 2
                    
        wb.save(filename)
        print(f"Combined results saved to {filename}")